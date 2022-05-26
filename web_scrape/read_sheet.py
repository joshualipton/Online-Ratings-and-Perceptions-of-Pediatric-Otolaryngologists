from openpyxl import load_workbook
import physician

# File and sheet being used
FILE_PATH = 'Scraping_Sheet_Short.xlsx'
SHEET_NAME = 'Sheet 1 - American Head and Nec'

# Creates a workbook to read from
WB = load_workbook(filename = FILE_PATH)
SHEET = WB[SHEET_NAME]

# Create Physician class that stores a dictionary for every row so that 
    # I can have a physician instance for each row
    # Maybe store each physician link in a dict bc each link has diff tags and stuff
# Includes ID, FN, LN, and list of links
# Store each physician in a big list to iterate over for my web scrape
# After scrape add all essential stuff to the object like ratings and comments
# Go through this list and add it to an excel file

# Modules: 
# read_sheet
# physician class
# web_scrape
# write_to_excel_sheet

# - Open link
# - First Name
# - Last Name
# - Residency
# - Put in tier
# - Source
# - Get comments
# - Date of rating
# - Rating
# - Indicate whether it was helpful or not if there
# - use datetime import for date object

def get_physicians() -> list[physician.Physician]:
    '''
    Gets all physicians from excel file
    And sets ID, first name, last name, and links
    Returns a physician object
    '''
    list_of_physicians = [] # Contains all physicians

    for line_number in range(2, SHEET.max_row):
        p = physician.Physician()
        p.set_id(SHEET['A' + str(line_number)].value)
        p.set_first_name((SHEET['B' + str(line_number)].value).strip())
        p.set_last_name((SHEET['C' + str(line_number)].value).strip())
        p.set_links(('Healthgrades', SHEET['D' + str(line_number)].value).strip())
        p.set_links(('Vitals', SHEET['E' + str(line_number)].value).strip())
        p.set_links(('RateMDs', SHEET['F' + str(line_number)].value).strip())
        p.set_links(('Yelp', SHEET['G' + str(line_number)].value).strip())
        p.set_residency((SHEET['H' + str(line_number)].value).strip())
        print(p.get_residency())

        list_of_physicians.append(p) # Adds physician to list

    return list_of_physicians